import openpyxl, json, os, sys, datetime

EXCEL = sys.argv[1] if len(sys.argv) > 1 else "/sessions/tender-amazing-curie/mnt/uploads/Sales YoY .3.xlsx"
wb = openpyxl.load_workbook(EXCEL, read_only=True, data_only=True)

# Helper: pick first existing sheet name
def get_sheet(*names):
    for n in names:
        if n in wb.sheetnames: return wb[n]
    raise KeyError(f"None found: {names}")

def s(v): return '' if v is None else str(v).strip()
def flt(v):
    try: return float(v)
    except: return 0.0
def fi(v): return int(round(flt(v)))

# ── 1. CUOTA JUL ─────────────────────────────────────────────────────────
# header row 2 (idx 2), data from row 3 (idx 3)
sh = get_sheet('1.Cuota Mes', 'Cuota Jul')
rows = list(sh.iter_rows(values_only=True))
cuota_jul = []
for r in rows[3:]:
    reg = s(r[0]); clv = s(r[2])
    if not reg or not clv: continue
    alc = flt(r[6]); alc = round(alc*100,1) if alc < 2 else round(alc,1)
    cuota_jul.append({'region':reg,'sub_region':s(r[1]),'clave':clv,'tienda':s(r[3]),
        'cuota':fi(r[4]),'ventas':fi(r[5]),'proy':0,'alcance':alc,
        'oh_total':0,'oh_cases':0,'oh_cables':0,'oh_chargers':0,'oh_liquid':0,'oh_micas':0,
        'resurtido':fi(r[8]),'faltante':fi(r[9]),'riesgo':s(r[10]) or 'Sin riesgo'})
print(f"Cuota Jul: {len(cuota_jul)} tiendas")

# ── 2. PROYECCION TIENDA → proy por clave ────────────────────────────────
# header row 2, data from row 3: 0=region,3=clave,4=tienda,5=alpha_actual,8=alpha_proy
pm = {}
try:
    sh_pt = get_sheet('2.Proyección Tienda', 'Proyección Tienda')
    for r in sh_pt.iter_rows(values_only=True, min_row=4):
        clv = s(r[3])
        if clv: pm[clv] = round(flt(r[8]),1)   # Alphacomm proy. cierre
except: pass
if not pm:  # fallback: AR sheet col 10
    for r in get_sheet('3.AR','1.AR').iter_rows(values_only=True, min_row=4):
        clv = s(r[3])
        if clv: pm[clv] = round(flt(r[10]),1)
for c in cuota_jul:
    c['proy'] = pm.get(c['clave'], c['ventas'])
print(f"Proy: {len(pm)} tiendas mapeadas")

# ── 3. 3.AR → ar_clean ───────────────────────────────────────────────────
# header row 2: 0=region,2=gerente,3=clave,4=tienda,5=ventas_alpha,8=equipo,10=alpha_proy,11=equipo_proy
ar_clean = []
for r in get_sheet('3.AR','1.AR').iter_rows(values_only=True, min_row=4):
    reg = s(r[0]); clv = s(r[3])
    if not reg or not clv: continue
    # proy: prefer pm dict (2.Proyección Tienda) over col K which may have uncached formulas
    proy_val = pm.get(clv, round(flt(r[10]), 1))
    ar_clean.append({'region':reg,'gerente':s(r[2]),'clave':clv,'tienda':s(r[4]),
        'alphacomm': fi(r[5]),   # Ventas Alphacomm del período actual
        'equipo':    fi(r[8]),   # Venta de equipo del período actual
        'proy':      proy_val})  # Alphacomm proy. cierre (desde 2.Proyección Tienda)
print(f"1.AR: {len(ar_clean)} registros")

# ── 4. OH DETALLE → oh_tienda, oh_productos, _rMap ───────────────────────
# Columns: 0=tienda,1=clave,2=regional,3=región,4=codigo(SKU),5=cat_alpha,
#          6=descripcion,7=modelo,8=subinv(numeric),10=cantidad
# Cols 12-14 from OH + Data items sheet provide descriptive names

# Load Data items lookup: codigo/sku → {desc, cat, subcat, clase}
try:
    sh_di = get_sheet('Data items')
    di_map = {}
    for dr in sh_di.iter_rows(values_only=True, min_row=3):
        sku  = s(dr[1])   # SKU Retail
        item = s(dr[2])   # ITEM code
        entry = {'desc':   s(dr[4]) or '',   # Descripción oficial
                 'cat':    s(dr[5]) or '',   # Categoría
                 'subcat': s(dr[6]) or '',   # Sub Categoría (descriptiva)
                 'clase':  s(dr[7]) or ''}  # Clase (descriptiva, ej. "6 ft")
        if sku:  di_map[sku]  = entry
        if item: di_map[item] = entry
    print(f"   Data items: {len(di_map)} entradas cargadas")
except Exception as e:
    di_map = {}
    print(f"   Data items: no encontrada ({e})")

sh4 = get_sheet('5.OH detalle', 'OH detalle')
t_map = {}   # clave → {clave,tienda,region,cantidad,cats:{}}
p_map = {}   # codigo → {producto,sub_categoria,clase,categoria,cantidad,by_region,codigo}
r_map = {}   # sub_cat|clase|cat → {sub_categoria,clase,categoria,cantidad,by_region}
reg_set4 = {}

for r in sh4.iter_rows(values_only=True, min_row=4):
    reg    = s(r[3]); clv = s(r[1]); tda = s(r[0])
    codigo = s(r[4])        # SKU / item code (N.XXXXXXX)
    qty    = fi(r[10])
    if not clv or qty <= 0: continue
    reg_set4[reg] = 1

    # Look up Data items for canonical/descriptive names
    di = di_map.get(codigo, {})
    cat_alpha = di.get('cat') or s(r[5]) or 'Accesories'
    prod      = di.get('desc') or s(r[6]) or cat_alpha
    sub_cat   = di.get('subcat') or s(r[7]) or cat_alpha
    clase     = di.get('clase') or s(r[8]) or 'General'

    if clv not in t_map:
        t_map[clv] = {'clave':clv,'tienda':tda,'region':reg,'gerente':'','cantidad':0,'cats':{}}
    t_map[clv]['cantidad'] += qty
    t_map[clv]['cats'][cat_alpha] = t_map[clv]['cats'].get(cat_alpha, 0) + qty

    # p_map: one entry per unique SKU/product
    pk = codigo
    if pk not in p_map:
        p_map[pk] = {'producto':prod,'sub_categoria':sub_cat,'clase':clase,
                     'categoria':cat_alpha,'cantidad':0,'by_region':{},'codigo':codigo}
    p_map[pk]['cantidad'] += qty
    p_map[pk]['by_region'][reg] = p_map[pk]['by_region'].get(reg, 0) + qty

    # r_map: descriptive sub_cat|clase pivot (all human-readable now)
    sk = sub_cat + '|' + clase + '|' + cat_alpha
    if sk not in r_map:
        r_map[sk] = {'sub_categoria':sub_cat,'clase':clase,'categoria':cat_alpha,
                     'cantidad':0,'by_region':{}}
    r_map[sk]['cantidad'] += qty
    r_map[sk]['by_region'][reg] = r_map[sk]['by_region'].get(reg, 0) + qty

oh_tienda = sorted(t_map.values(), key=lambda x: x['clave'])
oh_productos = sorted(p_map.values(), key=lambda x: -x['cantidad'])
cat_totals = {}
for p in oh_productos:
    cat_totals[p['categoria']] = cat_totals.get(p['categoria'],0) + p['cantidad']
oh_regiones = sorted(reg_set4.keys())

# Patch OH into cuota_jul
def mapcat(cat, qty, target):
    cl = cat.lower()
    if 'case' in cl: target['oh_cases'] += qty
    elif 'cable' in cl: target['oh_cables'] += qty
    elif 'charg' in cl or 'cargad' in cl: target['oh_chargers'] += qty
    elif 'liquid' in cl: target['oh_liquid'] += qty
    elif 'mica' in cl or 'protec' in cl: target['oh_micas'] += qty

oh_map = {t['clave']: t for t in oh_tienda}
for c in cuota_jul:
    oh = oh_map.get(c['clave'])
    if not oh: continue
    c['oh_total'] = oh['cantidad']
    for cat, qty in oh['cats'].items(): mapcat(cat, qty, c)

print(f"OH: {len(oh_tienda)} tiendas, {len(oh_productos)} productos, {len(oh_regiones)} regiones")

# ── 5. CUOTA LG-CASES JUL ────────────────────────────────────────────────
sh3 = get_sheet('4.Cuota LG-Cases', 'Cuota LG-Cases Jul')
rows3 = list(sh3.iter_rows(values_only=True))

def alc_pct(v):
    f = flt(v); return round(f*100,1) if f < 2 else round(f,1)

# Region summary: rows 8-13 (idx), header at idx 7
lg_region = []
for r in rows3[8:15]:
    reg = s(r[0])
    if not reg or reg.upper() == 'TOTAL': continue
    lg_region.append({'region':reg,
        'lg_v':fi(r[3]),'lg_cuota':round(flt(r[5]),1),'lg_alc':alc_pct(r[6]),
        'cases_v':fi(r[7]),'cases_cuota':round(flt(r[9]),1),'cases_alc':alc_pct(r[10])})

# Totals row 14
tot = rows3[14] if len(rows3) > 14 else [0]*11
lg_totals = {'lg_v':fi(tot[3]),'lg_cuota':round(flt(tot[5]),1),
             'cases_v':fi(tot[7]),'cases_cuota':round(flt(tot[9]),1)}

# Tienda detail: header at idx 17, data from 18
lg_tienda = []
for r in rows3[18:]:
    reg = s(r[0]); clv = s(r[1])
    if not clv: continue
    lg_tienda.append({'region':reg,'clave':clv,'tienda':s(r[2]),
        'lg_v':fi(r[3]),'lg_cuota':round(flt(r[5]),1),'lg_alc':alc_pct(r[6]),
        'cases_v':fi(r[7]),'cases_cuota':round(flt(r[9]),1),'cases_alc':alc_pct(r[10]),
        'oh_lg':0,'oh_cases':0})

# Patch OH into lg_tienda
for r in lg_tienda:
    oh = oh_map.get(r['clave'])
    if not oh: continue
    for cat, qty in oh['cats'].items():
        cl = cat.lower()
        if 'liquid' in cl: r['oh_lg'] += qty
        if 'case' in cl: r['oh_cases'] += qty

print(f"LG-Cases: {len(lg_region)} regiones, {len(lg_tienda)} tiendas")

# ── 6. 2.BASE (historical) ────────────────────────────────────────────────
MO = {'Jan':1,'Feb':2,'Mar':3,'Apr':4,'May':5,'Jun':6,'Jul':7,'Aug':8,'Sep':9,'Oct':10,'Nov':11,'Dec':12}
bm = {}; bRegSet = {}; bCatSet = {}; count = 0
# Acumuladores por tienda/clave para LG y Cases (leídos directo de 6.Base)
lg_by_clave   = {}  # clave → {mes|año: units_liquid}
cs_by_clave   = {}  # clave → {mes|año: units_cases}
sh5 = get_sheet('6.Base', '2.Base')
for i, r in enumerate(sh5.iter_rows(values_only=True, min_row=4)):
    mes = s(r[8]); units = fi(r[7])
    if not mes or units <= 0: continue
    año = fi(r[9]) or 2025
    key = mes+'|'+str(año)
    reg = s(r[0]); cat = s(r[10]); clv = s(r[3])
    if key not in bm:
        bm[key] = {'mes':mes,'año':año,'total':0,'by_region':{},'by_cat':{},'by_rc':{}}
    bm[key]['total'] += units
    if reg:
        bm[key]['by_region'][reg] = bm[key]['by_region'].get(reg,0) + units
        bRegSet[reg] = 1
    if cat:
        bm[key]['by_cat'][cat] = bm[key]['by_cat'].get(cat,0) + units
        bCatSet[cat] = 1
    if reg and cat:
        rc = reg+'|'+cat
        bm[key]['by_rc'][rc] = bm[key]['by_rc'].get(rc,0) + units
    # Acumular LG y Cases por clave para parchear lg_tienda después
    if clv:
        cl = cat.lower()
        if 'liquid' in cl:
            lg_by_clave.setdefault(clv, {})
            lg_by_clave[clv][key] = lg_by_clave[clv].get(key, 0) + units
        elif 'case' in cl:
            cs_by_clave.setdefault(clv, {})
            cs_by_clave[clv][key] = cs_by_clave[clv].get(key, 0) + units
    count += 1
    if count % 20000 == 0: print(f"  2.Base: {count} filas…")

sorted_keys = sorted(bm.keys(), key=lambda k:(int(k.split('|')[1]), MO.get(k.split('|')[0],0)))
base_data = {'byMonth': bm, 'sorted': sorted_keys, 'cats': sorted(bCatSet.keys())}
base_regiones = sorted(bRegSet.keys())
print(f"2.Base: {count} filas, {len(bm)} meses")

# ── 7. TOTALES + MES CUOTA ACTUAL ─────────────────────────────────────────
regiones = sorted(set(r['region'] for r in cuota_jul if r['region']))
proy_total = sum(r['proy'] for r in cuota_jul)

# Detectar mes de cuota: el mes siguiente al último mes completo (>= 5000 uds)
mo_list = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
complete_keys = [k for k in sorted_keys if bm[k]['total'] >= 5000]
if complete_keys:
    lmo, lyr = complete_keys[-1].split('|')
    idx = mo_list.index(lmo)
    cuota_mes = mo_list[idx+1] if idx < 11 else 'Jan'
    cuota_año = int(lyr) + (1 if idx == 11 else 0)
else:
    cuota_mes, cuota_año = 'Aug', 2026
print(f"Cuota mes detectado (raw): {cuota_mes}|{cuota_año}")
# No avanzar más allá del mes actual del calendario
_today = datetime.date.today()
_cur_mo = mo_list[_today.month - 1]
_cur_yr = _today.year
if cuota_año > _cur_yr or (cuota_año == _cur_yr and MO[cuota_mes] > _today.month):
    cuota_mes, cuota_año = _cur_mo, _cur_yr
    print(f"Ajustado al mes actual: {cuota_mes}|{cuota_año}")
print(f"Cuota mes final: {cuota_mes}|{cuota_año}")

# ── 7a. PARCHAR lg_tienda CON VENTAS REALES DE 6.BASE ─────────────────────
# Las fórmulas SUMIFS de 4.Cuota LG-Cases pueden tener caché desactualizado.
# Se reemplazan lg_v y cases_v con la suma directa de 6.Base por clave/mes.
# El mes activo de ventas = cuota_mes (el mes que se está midiendo en 3.AR).
_cur_key = cuota_mes + '|' + str(cuota_año)

# Verificar si 6.Base tiene desglose COMPLETO de LG/Cases para el mes actual.
# Comparar contra el total cacheado del Excel (lg_totals viene de 4.Cuota LG-Cases).
# Si 6.Base tiene < 80% de lo que indica el Excel, los datos son parciales → fallback.
total_lg_base  = sum(v for d in lg_by_clave.values() for k, v in d.items() if k == _cur_key)
total_cs_base  = sum(v for d in cs_by_clave.values() for k, v in d.items() if k == _cur_key)
excel_lg_ref   = lg_totals.get('lg_v', 0) + lg_totals.get('cases_v', 0)
base_lg_total  = total_lg_base + total_cs_base
# Usar 6.Base solo si cubre ≥80% de lo que el Excel indica (datos completos post-fix)
base_has_cats  = base_lg_total > 0 and (excel_lg_ref == 0 or base_lg_total >= excel_lg_ref * 0.8)

lg_patched = cs_patched = 0
if base_has_cats:
    # 6.Base tiene categorías reales → usarlas (caso normal post-fix)
    for t in lg_tienda:
        clv = t['clave']
        new_lg = lg_by_clave.get(clv, {}).get(_cur_key, 0)
        new_cs = cs_by_clave.get(clv, {}).get(_cur_key, 0)
        if new_lg != t['lg_v']: t['lg_v'] = new_lg; lg_patched += 1
        if new_cs != t['cases_v']: t['cases_v'] = new_cs; cs_patched += 1
        t['lg_alc']    = round(new_lg / t['lg_cuota'] * 100, 1) if t['lg_cuota'] else 0.0
        t['cases_alc'] = round(new_cs / t['cases_cuota'] * 100, 1) if t['cases_cuota'] else 0.0
    # Recalcular totales y region summary
    lg_totals = {
        'lg_v':        sum(t['lg_v'] for t in lg_tienda),
        'lg_cuota':    lg_totals.get('lg_cuota', 0),
        'cases_v':     sum(t['cases_v'] for t in lg_tienda),
        'cases_cuota': lg_totals.get('cases_cuota', 0)
    }
    for reg_row in lg_region:
        reg = reg_row['region']
        reg_row['lg_v']      = sum(t['lg_v']    for t in lg_tienda if t['region'] == reg)
        reg_row['cases_v']   = sum(t['cases_v'] for t in lg_tienda if t['region'] == reg)
        reg_row['lg_alc']    = round(reg_row['lg_v']    / reg_row['lg_cuota']    * 100, 1) if reg_row['lg_cuota']    else 0.0
        reg_row['cases_alc'] = round(reg_row['cases_v'] / reg_row['cases_cuota'] * 100, 1) if reg_row['cases_cuota'] else 0.0
    print(f"LG tienda parche (6.Base): LG={lg_totals['lg_v']} Cases={lg_totals['cases_v']}")
else:
    # 6.Base no tiene desglose de categorías para el mes actual (extract_primemx.py pendiente de fix)
    # Conservar valores cacheados de 4.Cuota LG-Cases (la hoja Excel los tiene aunque stale)
    print(f"LG tienda: 6.Base sin categorías para {_cur_key} → conservando caché de 4.Cuota LG-Cases "
          f"(LG={lg_totals.get('lg_v',0)} Cases={lg_totals.get('cases_v',0)})")

# ── 7b. PARCHAR MES ACTUAL CON DATOS DE 3.AR ──────────────────────────────
# 3.AR refleja las ventas del cuota_mes (mes en curso que se está midiendo).
# Solo se reemplaza si 3.AR tiene más unidades que 6.Base (datos más completos).
cur_key = cuota_mes + '|' + str(cuota_año)

ar_total_cur = sum(r['alphacomm'] for r in ar_clean)
ar_by_region = {}
for r in ar_clean:
    ar_by_region[r['region']] = ar_by_region.get(r['region'], 0) + r['alphacomm']

if cur_key not in bm:
    bm[cur_key] = {'mes': cuota_mes, 'año': cuota_año,
                   'total': 0, 'by_region': {}, 'by_cat': {}, 'by_rc': {}}
    sorted_keys = sorted(bm.keys(), key=lambda k:(int(k.split('|')[1]), MO.get(k.split('|')[0],0)))
    base_data['sorted'] = sorted_keys

old_total = bm[cur_key].get('total', 0)
if ar_total_cur > old_total:
    bm[cur_key]['total'] = ar_total_cur
    bm[cur_key]['by_region'] = ar_by_region
    print(f"Parche {cur_key}: {old_total} → {ar_total_cur} (fuente: 3.AR)")
else:
    print(f"Sin parche {cur_key}: 6.Base={old_total} >= AR={ar_total_cur} (6.Base más completo)")

# ── 8. EXPORTAR ───────────────────────────────────────────────────────────
data = {
    'cuota_mes': cuota_mes, 'cuota_año': cuota_año,
    'cuota_jul': cuota_jul, 'ar_clean': ar_clean,
    'lg_region': lg_region, 'lg_tienda': lg_tienda,
    'lg_totals': lg_totals, 'cat_totals': cat_totals,
    'regiones': regiones, 'proy_total': proy_total,
    'oh_tienda': oh_tienda, 'oh_productos': oh_productos,
    'oh_regiones': oh_regiones, 'rMap': r_map,
    'base_data': base_data, 'base_regiones': base_regiones,
    'generated_at': datetime.datetime.now().strftime('%Y-%m-%dT%H:%M:%S')
}

out = sys.argv[2] if len(sys.argv) > 2 else '/sessions/brave-zealous-einstein/mnt/2.Prime/data.json'
with open(out,'w',encoding='utf-8') as f:
    json.dump(data, f, ensure_ascii=False, separators=(',',':'))
sz = os.path.getsize(out)
print(f"\n✅ data.json: {sz/1024:.0f} KB  ({sz/1024/1024:.2f} MB)")
