"""
extract_primemx.py  —  Versión 2 (memory-efficient)
Uso: python3 extract_primemx.py <PrimeMX.xlsx> <SalesYoY.xlsx>

Actualiza directamente el ZIP sin cargar el workbook completo:
  • 3.AR       ← Detalle por Tienda  (reemplaza datos, conserva fórmulas K-M)
  • 6.Base     ← Detalle por Tienda  (ventas totales por tienda, reemplaza mes actual)
  • 1.Cuota Mes← Detalle por Tienda  (actualiza col F Ventas mes con alphacomm real)

El resultado se guarda como "Sales YoY .{n+1}.xlsx" (versión consecutiva).
Imprime la ruta del archivo de salida para que el caller pueda usarla.
"""
import sys, os, re, zipfile, shutil
from datetime import datetime
from xml.etree import ElementTree as ET
import openpyxl

# ── Argumentos ───────────────────────────────────────────────────────────────
if len(sys.argv) < 3:
    print("Uso: python3 extract_primemx.py <PrimeMX.xlsx> <SalesYoY.xlsx>")
    sys.exit(1)
PRIMEMX, SALES = sys.argv[1], sys.argv[2]

# ── Calcular número de versión siguiente ─────────────────────────────────────
folder = os.path.dirname(SALES)
existing = [f for f in os.listdir(folder) if re.match(r'Sales YoY \.\d+\.xlsx$', f)]
versions = [int(re.search(r'\.(\d+)\.xlsx$', f).group(1)) for f in existing]
next_ver  = max(versions) + 1 if versions else 1
SALES_OUT = os.path.join(folder, f"Sales YoY .{next_ver}.xlsx")
print(f"📄 Salida: Sales YoY .{next_ver}.xlsx (anterior: .{max(versions) if versions else '?'}.xlsx)")

# ── Fecha desde nombre del archivo ───────────────────────────────────────────
m = re.search(r'(\d{4})-(\d{2})-(\d{2})', os.path.basename(PRIMEMX))
if not m:
    raise ValueError(f"El nombre debe contener YYYY-MM-DD: {os.path.basename(PRIMEMX)}")
year, month_num, day = int(m[1]), int(m[2]), int(m[3])
MONTHS = {1:'Jan',2:'Feb',3:'Mar',4:'Apr',5:'May',6:'Jun',
          7:'Jul',8:'Aug',9:'Sep',10:'Oct',11:'Nov',12:'Dec'}
month_str = MONTHS[month_num]
cut_date  = datetime(year, month_num, day)
print(f"📅 Fecha de corte: {cut_date.strftime('%d %b %Y')} → '{month_str}' {year}")

# ── [1/5] Leer PrimeMX (archivo pequeño, openpyxl normal) ───────────────────
print("\n[1/5] Leyendo PrimeMX...")
wb_pm = openpyxl.load_workbook(PRIMEMX, read_only=True, data_only=True)
dt_data  = [r[:10] for r in wb_pm['Detalle por Tienda'].iter_rows(values_only=True, min_row=5) if r[0]]
dcm_data = [r[:8]  for r in wb_pm['Detalle Completo Modelos'].iter_rows(values_only=True, min_row=5) if r[0]]
wb_pm.close()
print(f"   {len(dt_data)} tiendas · {len(dcm_data)} líneas de modelos")

# Índice clave → alphacomm para actualizar 1.Cuota Mes
dt_by_clave = {str(r[3]).strip(): int(r[5] or 0) for r in dt_data if r[3]}

# ── [2/5] Leer Sales YoY en modo streaming (bajo consumo de memoria) ─────────
print("\n[2/5] Leyendo Sales YoY en modo streaming...")
wb_ro = openpyxl.load_workbook(SALES, read_only=True, data_only=True)

sku_cat = {}
for r in wb_ro['Data items'].iter_rows(values_only=True, min_row=3):
    sku = str(r[1]).strip() if r[1] else ''
    if sku:
        sku_cat[sku] = {'cat': r[5], 'sub': r[6], 'clase': r[7],
                        'costo': float(r[8]) if r[8] else 0.0}
print(f"   Catálogo SKU: {len(sku_cat)} items")

existing_base, skipped = [], 0
for r in wb_ro['6.Base'].iter_rows(values_only=True, min_row=4):
    if not r[0]:
        continue
    if r[8] == month_str and r[9] == year:
        skipped += 1
    else:
        existing_base.append(r)
print(f"   6.Base: {len(existing_base)} filas previas · {skipped} de {month_str} {year} reemplazadas")

# Leer 1.Cuota Mes para actualización de ventas
cuota_rows = []
if '1.Cuota Mes' in wb_ro.sheetnames:
    for r in wb_ro['1.Cuota Mes'].iter_rows(values_only=True, min_row=4):
        if r[0]:
            cuota_rows.append(list(r))
print(f"   1.Cuota Mes: {len(cuota_rows)} tiendas leídas")

# Índice prefijo ciudad → sub_región (para asignar a tiendas nuevas)
prefix_to_sub = {}
for r in cuota_rows:
    clave_c = str(r[2]).strip() if r[2] else ''
    if r[1] and len(clave_c) >= 9:
        prefix_to_sub.setdefault(clave_c[6:9], r[1])
wb_ro.close()

# ── Construir nuevas filas para 6.Base (desde DCM = ventas por modelo/SKU) ───
# Se usa dcm_data (Detalle Completo Modelos) en lugar de dt_data para preservar
# el desglose por categoría (Liquid, Cases iPhone, Cables, etc.) que usan las
# fórmulas SUMIFS de la hoja 4.Cuota LG-Cases.
# Columnas DCM esperadas: [0]Región [1]Director [2]Gerente [3]Clave [4]Nombre
#                         [5]SKU/Modelo [6]Unidades [7]Costo unitario
new_base = []
dt_clave_set = {str(r[3]).strip() for r in dt_data if r[3]}  # para validación

for r in dcm_data:
    region   = r[0]
    director = r[1]
    gerente  = r[2]
    clave    = str(r[3]).strip() if r[3] else ''
    nombre   = r[4]
    sku      = str(r[5]).strip() if len(r) > 5 and r[5] else ''
    qty      = int(round(float(r[6]))) if len(r) > 6 and r[6] else 0
    costo_u  = float(r[7]) if len(r) > 7 and r[7] else 0.0
    if not clave or qty <= 0:
        continue
    info  = sku_cat.get(sku, {})
    cat   = (info.get('cat') or 'Accesories') if info else 'Accesories'
    sub   = info.get('sub') or ''
    clase = info.get('clase') or ''
    if info and info.get('costo'):
        costo_u = info['costo']
    new_base.append((region, director, gerente, clave, nombre,
                     sku, '', qty,
                     month_str, year,
                     cat, sub, clase, costo_u, round(qty * costo_u, 2), None))

# Fallback: tiendas que están en DT pero sin detalle en DCM → agregar como Alphacomm
dcm_claves = {str(r[3]).strip() for r in dcm_data if r[3] and (int(round(float(r[6]))) if len(r) > 6 and r[6] else 0) > 0}
for r in dt_data:
    clave = str(r[3]).strip() if r[3] else ''
    alphacomm = int(r[5] or 0) if r[5] else 0
    if not clave or alphacomm <= 0 or clave in dcm_claves:
        continue
    new_base.append((r[0], r[1], r[2], clave, r[4],
                     'General', 'ALPHA', alphacomm,
                     month_str, year,
                     'Alphacomm', '', '', None, None, None))

# Desglose de categorías para validación
from collections import Counter
cat_check = Counter(row[10] for row in new_base)
all_base = existing_base + new_base
print(f"   Total 6.Base: {len(all_base)} filas ({len(new_base)} filas nuevas de {month_str} {year})")
print(f"   Categorías: {dict(cat_check)}")

# ── Helpers XML ───────────────────────────────────────────────────────────────
EPOCH = datetime(1899, 12, 30)

def col_ltr(n):
    s = ''
    while n > 0:
        n, r = divmod(n-1, 26)
        s = chr(65+r) + s
    return s

def esc(v):
    return str(v).replace('&','&amp;').replace('<','&lt;').replace('>','&gt;')

def cx(col, row, val):
    ref = f"{col_ltr(col)}{row}"
    if val is None:           return f'<c r="{ref}"/>'
    if isinstance(val, bool): return f'<c r="{ref}" t="b"><v>{int(val)}</v></c>'
    if isinstance(val, datetime):
        return f'<c r="{ref}" s="1"><v>{(val-EPOCH).days}</v></c>'
    if isinstance(val, (int, float)): return f'<c r="{ref}"><v>{val}</v></c>'
    return f'<c r="{ref}" t="inlineStr"><is><t>{esc(val)}</t></is></c>'

# ── [3/5] Obtener rutas de hojas dentro del ZIP ───────────────────────────────
def get_sheet_paths(zip_obj):
    PKG = 'http://schemas.openxmlformats.org/package/2006/relationships'
    WB  = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
    REL = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
    wb_xml   = ET.fromstring(zip_obj.read('xl/workbook.xml'))
    rels_xml = ET.fromstring(zip_obj.read('xl/_rels/workbook.xml.rels'))
    rid2tgt  = {r.get('Id'): r.get('Target') for r in rels_xml.findall(f'{{{PKG}}}Relationship')}
    result   = {}
    for s in wb_xml.findall(f'.//{{{WB}}}sheet'):
        rid = s.get(f'{{{REL}}}id')
        tgt = rid2tgt.get(rid, '')
        result[s.get('name')] = ('xl/' + tgt) if not tgt.startswith('/') else tgt.lstrip('/')
    return result

print("\n[3/5] Leyendo estructura del ZIP y fórmulas de 3.AR y 1.Cuota Mes...")
with zipfile.ZipFile(SALES, 'r') as z:
    sheet_paths   = get_sheet_paths(z)
    ar_path       = sheet_paths.get('3.AR')
    base_path     = sheet_paths.get('6.Base')
    cuota_path    = sheet_paths.get('1.Cuota Mes')
    if not ar_path:
        raise ValueError("Hoja '3.AR' no encontrada en el archivo")
    ar_xml_raw    = z.read(ar_path).decode('utf-8')
    base_xml_raw  = z.read(base_path).decode('utf-8')
    cuota_xml_raw = z.read(cuota_path).decode('utf-8') if cuota_path else None

# Extraer fórmulas de K, L, M fila 4 de 3.AR como plantilla
tmpl_formulas = {}
for col_l in ('K', 'L', 'M'):
    fm = re.search(rf'<c r="{col_l}4"[^>]*>.*?<f>(.*?)</f>', ar_xml_raw, re.DOTALL)
    if fm:
        tmpl_formulas[col_l] = fm.group(1)
print(f"   Fórmulas 3.AR K/L/M: {list(tmpl_formulas.keys())}")

# Extraer fórmulas de G y J fila 4 de 1.Cuota Mes como plantilla
cuota_formulas = {}
if cuota_xml_raw:
    for col_l in ('G', 'J'):
        fm = re.search(rf'<c r="{col_l}4"[^>]*>.*?<f>(.*?)</f>', cuota_xml_raw, re.DOTALL)
        if fm:
            cuota_formulas[col_l] = fm.group(1)
    print(f"   Fórmulas 1.Cuota Mes G/J: {list(cuota_formulas.keys())}")

# ── [4/5] Construir nuevo XML para 3.AR ──────────────────────────────────────
print("\n[4/5] Construyendo XML actualizado...")

def adjust_formula(formula, new_row):
    """Reemplaza referencias de fila 4 → new_row en fórmulas relativas."""
    return re.sub(r'([A-Z]+)4(?!\d)', lambda m: m.group(1) + str(new_row), formula)

# Separar encabezado/pie de <sheetData>
sd_start = ar_xml_raw.find('<sheetData')
sd_end   = ar_xml_raw.find('</sheetData>') + len('</sheetData>')
ar_head  = ar_xml_raw[:sd_start]
ar_foot  = ar_xml_raw[sd_end:]

# Filas 1-3 originales
orig_ar_rows = re.findall(r'<row r="[123]".*?</row>', ar_xml_raw, re.DOTALL)
# Actualizar fecha en K2
orig_ar_rows = [re.sub(r'<c r="K2"[^>]*>.*?</c>', cx(11, 2, cut_date), row, flags=re.DOTALL)
                if 'r="2"' in row else row for row in orig_ar_rows]

ar_parts = ['<sheetData>'] + orig_ar_rows
for i, row_data in enumerate(dt_data):
    rn = 4 + i
    cells = ''.join(cx(j+1, rn, v) for j, v in enumerate(row_data))
    for col_l, formula in tmpl_formulas.items():
        col_n = {'K':11,'L':12,'M':13}[col_l]
        adj = adjust_formula(formula, rn)
        cells += f'<c r="{col_l}{rn}"><f>{adj}</f></c>'
    ar_parts.append(f'<row r="{rn}">{cells}</row>')
ar_parts.append('</sheetData>')
new_ar_xml = ar_head + ''.join(ar_parts) + ar_foot

# ── Construir nuevo XML para 6.Base ──────────────────────────────────────────
sd_start_b = base_xml_raw.find('<sheetData')
sd_end_b   = base_xml_raw.find('</sheetData>') + len('</sheetData>')
base_head  = base_xml_raw[:sd_start_b]
base_foot  = base_xml_raw[sd_end_b:]

orig_base_rows = re.findall(r'<row r="[123]".*?</row>', base_xml_raw, re.DOTALL)
base_parts = ['<sheetData>'] + orig_base_rows
for i, row_data in enumerate(all_base):
    rn = 4 + i
    cells = ''.join(cx(j+1, rn, v) for j, v in enumerate(row_data))
    base_parts.append(f'<row r="{rn}">{cells}</row>')
base_parts.append('</sheetData>')
new_base_xml = base_head + ''.join(base_parts) + base_foot

# ── Construir nuevo XML para 1.Cuota Mes ─────────────────────────────────────
new_cuota_xml = None
if cuota_xml_raw and cuota_rows:
    sd_start_c = cuota_xml_raw.find('<sheetData')
    sd_end_c   = cuota_xml_raw.find('</sheetData>') + len('</sheetData>')
    cuota_head = cuota_xml_raw[:sd_start_c]
    cuota_foot = cuota_xml_raw[sd_end_c:]

    orig_cuota_rows = re.findall(r'<row r="[123]".*?</row>', cuota_xml_raw, re.DOTALL)
    cuota_parts = ['<sheetData>'] + orig_cuota_rows

    updated_cuota = 0
    cuota_claves_set = {str(r[2]).strip() for r in cuota_rows if r[2]}

    for i, row_data in enumerate(cuota_rows):
        rn = 4 + i
        clave     = str(row_data[2]).strip() if row_data[2] else ''
        cuota_val = float(row_data[4] or 0)
        # Stores not in DT = 0 ventas (no alphacomm sales that period)
        ventas_new = dt_by_clave.get(clave, 0)
        if clave in dt_by_clave:
            updated_cuota += 1

        cells = ''.join(cx(j+1, rn, row_data[j]) for j in range(5))  # A-E
        cells += cx(6, rn, ventas_new)  # F = Ventas mes actualizado

        if 'G' in cuota_formulas:  # G = Alcance
            cells += f'<c r="G{rn}"><f>{adjust_formula(cuota_formulas["G"], rn)}</f></c>'
        else:
            alcance = round(ventas_new / cuota_val, 10) if cuota_val > 0 else 0
            cells += cx(7, rn, alcance)

        for j in range(7, 9):  # H = OH, I = Resurtido
            cells += cx(j+1, rn, row_data[j] if j < len(row_data) else None)

        if 'J' in cuota_formulas:  # J = Faltante
            cells += f'<c r="J{rn}"><f>{adjust_formula(cuota_formulas["J"], rn)}</f></c>'
        else:
            faltante = max(0, int(cuota_val) - ventas_new)
            cells += cx(10, rn, faltante)

        # Columnas K en adelante — conservar valores originales (cols 11-16)
        for j in range(10, len(row_data)):
            cells += cx(j+1, rn, row_data[j])

        cuota_parts.append(f'<row r="{rn}">{cells}</row>')

    # Agregar tiendas que están en DT pero NO en 1.Cuota Mes
    new_stores_added = []
    for r in dt_data:
        clave     = str(r[3]).strip() if r[3] else ''
        alphacomm = int(r[5] or 0) if r[5] else 0
        if not clave or clave in cuota_claves_set or alphacomm <= 0:
            continue
        rn = 4 + len(cuota_rows) + len(new_stores_added)
        prefix    = clave[6:9] if len(clave) >= 9 else ''
        sub       = prefix_to_sub.get(prefix, '')
        oh        = int(r[8] or 0) if len(r) > 8 else 0
        cells = (cx(1, rn, r[0]) +       # A = Región
                 cx(2, rn, sub) +         # B = Sub_región (inferida)
                 cx(3, rn, clave) +       # C = Clave
                 cx(4, rn, r[4]) +        # D = Nombre
                 cx(5, rn, 0) +           # E = Cuota = 0 (pendiente asignar)
                 cx(6, rn, alphacomm) +   # F = Ventas
                 cx(7, rn, 0) +           # G = Alcance = 0
                 cx(8, rn, oh) +          # H = OH
                 cx(9, rn, None) +        # I = Resurtido
                 cx(10, rn, 0))           # J = Faltante
        cuota_parts.append(f'<row r="{rn}">{cells}</row>')
        new_stores_added.append(clave)

    cuota_parts.append('</sheetData>')
    new_cuota_xml = cuota_head + ''.join(cuota_parts) + cuota_foot

    # Actualizar <dimension> para que Excel y openpyxl reconozcan las nuevas filas
    if new_stores_added:
        last_row = 3 + len(cuota_rows) + len(new_stores_added)
        new_cuota_xml = re.sub(
            r'<dimension ref="[^"]*"/>',
            f'<dimension ref="A1:P{last_row}"/>',
            new_cuota_xml
        )

    print(f"   1.Cuota Mes: {updated_cuota}/{len(cuota_rows)} tiendas actualizadas · {len(new_stores_added)} nuevas añadidas: {new_stores_added}")

# ── [5/5] Escribir archivo actualizado (swapping solo las hojas cambiadas) ───
print("\n[5/5] Escribiendo archivo actualizado...")
tmp_path = SALES + '.tmp'
with zipfile.ZipFile(SALES, 'r') as z_in:
    with zipfile.ZipFile(tmp_path, 'w', compression=zipfile.ZIP_DEFLATED) as z_out:
        for item in z_in.infolist():
            if item.filename == ar_path:
                z_out.writestr(item, new_ar_xml.encode('utf-8'))
            elif item.filename == base_path:
                z_out.writestr(item, new_base_xml.encode('utf-8'))
            elif item.filename == cuota_path and new_cuota_xml:
                z_out.writestr(item, new_cuota_xml.encode('utf-8'))
            else:
                z_out.writestr(item, z_in.read(item.filename))

shutil.move(tmp_path, SALES_OUT)

print(f"\n🎉 Completado.")
print(f"   3.AR       : {len(dt_data)} tiendas · corte {cut_date.strftime('%d/%m/%Y')}")
print(f"   6.Base     : {len(all_base)} filas totales · {len(new_base)} tiendas nuevas de {month_str} {year}")
print(f"   1.Cuota Mes: {len(cuota_rows)} tiendas · ventas actualizadas desde DT")
print(f"   OUTPUT     : {SALES_OUT}")
